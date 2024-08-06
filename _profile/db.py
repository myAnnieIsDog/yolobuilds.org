# Standard Library modules
from datetime import datetime
import json
from pprint import pformat, pprint

# Third party modules
import django
from django.conf import settings
from django.core.wsgi import get_wsgi_application
from openpyxl import Workbook, load_workbook

# Django setting must be setup prior to referencing project modules.
application = get_wsgi_application()
settings.configure()
django.setup()
print(f'- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -')
print(f'settings: {settings}')
print(f'settings.DEBUG: %'.format(settings.DEBUG))

# Project modules
import models

print('READY!')
print(f'- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -')

def main():
    
    wb = load_workbook("db.xlsx")
    conf = {  
        # key = model
        # value = list of fields for the model
        #
        # These two limitations allow this function to utilize an Excel
        # file that does not match the models in order and/or naming:
        #   values must match '''xlsx''' file order. (OK to vary name)
        #   values must match '''model''' field name. (OK to mix order)
        
        "agency": ["id", "name", "full"],
        "certification": ["id", "agency", "agency_long", "license", "license_long"],
        "department": ["id", "name", "full"],
        "division": ["id", "prefix", "division", "full_division"],
        "profile": [
            "id",
            "user",
            "first",
            "last",
            "email",
            "phone",
            "address",
            "city",
            "state",
            "zip",
            "staff",
            "superadmin",
            "company",
            "agency",
            "department",
            "division",
            "reviewer",
            "inspector",
            "alt_contact",
            "alt_contact_email",
            "alt_contact_phone",
        ],
    }
    a = models.Agency()
    b = models.Certification()
    c = models.Department()
    d = models.Division()
    e = models.Profile
    mods = [a, b, c, d, e]
    handler(wb, conf, mods)



def handler(wb, conf:dict, app_models:list) -> None:
    # Prepare and respond to conversion
    data = {}
    for model, keys in conf.items():
        data.update({model: from_xlsx(wb[model], keys)})
    data.update({"Timestamp": datetime.now().strftime("%Y-%m-%d %H-%M-%S.%f %Z")})
    t = data["Timestamp"]
    print(f"xlsx read {t}")
    load = to_json_file(data, t)
    print(f"json written {datetime.now().strftime("%Y-%m-%d %H-%M-%S.%f %Z")}")
    to_sql(load, app_models)
    print(f"sql saved {datetime.now().strftime("%Y-%m-%d %H-%M-%S.%f %Z")}")


def from_xlsx(ws, keys, active=True):
    # Convert from xlsx to json
    list = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        dictionary = {}
        for h in keys:
            dictionary[h] = row[keys.index(h)]
        dictionary["active"] = active
        list.append(dictionary)
    return list


def to_json_file(data, t):
    load = json.dumps(data)
    # Save to long-term archive
    with open(f"data/db {t}.json", mode="w") as f:
        f.write(load)

    # Write to active database verifier
    with open(f"db.json", mode="w") as f:
        f.write(load)

    return load


def to_sql(load, app_models):
    print(load)
    print(app_models)


if __name__ == "__main__":
    main()
