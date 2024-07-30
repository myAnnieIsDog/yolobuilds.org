import json
from django.contrib.auth.models import User
from django.core import serializers
from django.db import IntegrityError

import openpyxl

from fees.models import Account, FeeType, PaymentMethod
from inspections.models import InspectionType, InspectionResult
from locations.models import District, Jurisdiction
from permits.models import Tag, Division, PermitStatus, PermitType, Sequence
from permits_bp.models import ApplicantRole, OwnerRole
from profiles.models import Profile, LicenseAgency, LicenseType, Agency, Department, Division, Staff
from reviews.models import ReviewType, ReviewStatus, CycleResult

##########################################################################
""" python manage.py runscript import """
##########################################################################

base = ".init/"

def run():
    fiscal()
    inspections()
    locations()
    profiles()
    # reviews()
    

def fiscal():
    wb = openpyxl.load_workbook(f'{base}fiscal.xlsx')
    ws = wb["accounts"]
    for i in range(2, 39):
        a = Account()
        a.fund = ws[f"B{i}"].value
        a.fund_label = ws[f"C{i}"].value
        a.share = ws[f"D{i}"].value
        a.unit = ws[f"E{i}"].value
        a.unit_label = ws[f"F{i}"].value
        a.unit_description = ws[f"G{i}"].value
        a.cost_center = ws[f"H{i}"].value
        a.gl_account = ws[f"I{i}"].value
        a.cams = ws[f"J{i}"].value
        a.infor_activity = ws[f"K{i}"].value
        a.infor_account = ws[f"L{i}"].value
        a.ledger = ws[f"M{i}"].value
        a.save()
    
    ws = wb["fee_types"]
    for i in range(2, 139):
        a = FeeType()
        # a.fee_account = ws[f"B{i}"].value
        a.fee_group = ws[f"C{i}"].value
        a.fee_type = ws[f"D{i}"].value
        a.policy = ws[f"E{i}"].value
        a.authorization = ws[f"F{i}"].value
        a.adopted = ws[f"G{i}"].value
        a.revised = ws[f"H{i}"].value
        a.expires = ws[f"I{i}"].value
        a.tier_base_qty = ws[f"J{i}"].value
        a.tier_base_fee = ws[f"K{i}"].value
        a.rate = ws[f"L{i}"].value
        a.units = ws[f"M{i}"].value
        a.rate_check = ws[f"N{i}"].value
        a.active = True
        a.deleted = False
        a.save()

    ws = wb["payment_methods"]
    for i in range(2, 8):
        a = PaymentMethod()
        a.method = ws[f"B{i}"].value
        a.policy = ws[f"C{i}"].value
        a.save()

def locations():
    wb = openpyxl.load_workbook(f'{base}locations.xlsx')
    ws = wb["districts"]
    for i in range(2, 46):
        a = District()
        a.district_type = ws[f"B{i}"].value
        a.district = ws[f"C{i}"].value
        a.description = ws[f"D{i}"].value
        a.save()

    wb = openpyxl.load_workbook(f'{base}locations.xlsx')
    ws = wb["jurisdictions"]
    for i in range(2, 6):
        a = Jurisdiction()
        a.jurisdiction = ws[f"B{i}"].value
        a.save()

def inspections():
    wb = openpyxl.load_workbook(f'{base}inspections.xlsx')
    ws = wb["insp_type"]
    for i in range(2, 68):
        a = InspectionType()
        a.inspection_type = ws[f"B{i}"].value
        a.save()

    wb = openpyxl.load_workbook(f'{base}inspections.xlsx')
    ws = wb["insp_result"]
    for i in range(2, 8):
        a = InspectionResult()
        a.result = ws[f"B{i}"].value
        a.requirements = ws[f"C{i}"].value
        a.save()

def profiles():
    wb = openpyxl.load_workbook(f'{base}profiles.xlsx')
    ws = wb["agency"]
    for i in range(2, 14):
        a = Agency()
        a.agency = ws[f"B{i}"].value
        a.full_agency = ws[f"C{i}"].value
        a.save()
    
    wb = openpyxl.load_workbook(f'{base}profiles.xlsx')
    ws = wb["department"]
    for i in range(2, 18):
        a = Department()
        a.dept_code = ws[f"B{i}"].value
        a.department = ws[f"C{i}"].value
        a.save()
    
    wb = openpyxl.load_workbook(f'{base}profiles.xlsx')
    ws = wb["division"]
    for i in range(2, 9):
        a = Division()
        a.prefix = ws[f"B{i}"].value
        a.division = ws[f"C{i}"].value
        a.full_division = ws[f"D{i}"].value
        a.save()
    
    wb = openpyxl.load_workbook(f'{base}profiles.xlsx')
    ws = wb["profile"]
    for i in range(2, 31):
        a = Profile()
        # a.user = ws[f"B{i}"].value
        a.first = ws[f"C{i}"].value
        a.last = ws[f"D{i}"].value
        a.company = ws[f"E{i}"].value
        a.email = ws[f"F{i}"].value
        a.phone = ws[f"G{i}"].value
        a.address = ws[f"H{i}"].value
        a.city = ws[f"I{i}"].value
        a.state = ws[f"J{i}"].value
        a.zip = ws[f"K{i}"].value
        a.save()
    
    wb = openpyxl.load_workbook(f'{base}profiles.xlsx')
    ws = wb["profile"]
    for i in range(2, 31):
        a = User()
        a.username = ws[f"B{i}"].value
        a.first_name = ws[f"C{i}"].value
        a.last_name = ws[f"D{i}"].value
        a.email = ws[f"F{i}"].value
        a.is_active = ws[f"L{i}"].value
        a.is_staff = ws[f"M{i}"].value
        a.is_superuser = ws[f"N{i}"].value
        a.save()

    wb = openpyxl.load_workbook(f'{base}profiles.xlsx')
    ws = wb["licenser"]
    for i in range(2, 6):
        a = LicenseAgency()
        a.agency = ws[f"B{i}"].value
        a.agency_long = ws[f"C{i}"].value
        a.save()

    wb = openpyxl.load_workbook(f'{base}profiles.xlsx')
    ws = wb["license_type"]
    for i in range(2, 54):
        a = LicenseType()
        a.license_short = ws[f"B{i}"].value
        a.license_long = ws[f"C{i}"].value
        a.save()

if __name__ == '__main__':
    run()
# ##########################################################################
# """ End File """
# ##########################################################################