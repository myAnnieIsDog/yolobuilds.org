""" 
Django Extensions looks in project/Scripts for a package with __init__.py then
opens the file matching the script name and runs the run() function.

python manage.py runscript db 
    --script-args ___
    --chdir /project/app
    --dir-policy (none=current; each=script directory; root=base directory)
    --continue-on-error
    --traceback
    --no-traceback

https://django-extensions.readthedocs.io/en/latest/runscript.html
similar to: python manage.py shell 
"""

# Standard Library modules
import argparse
import csv  # https://docs.python.org/3/library/csv.html
from datetime import datetime as dt
import json
import os
from pprint import pformat, pprint
import random
import time

# Third-party modules
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "settings")
from django.contrib.auth.models import User
from django.db.utils import IntegrityError
from openpyxl import Workbook, load_workbook

# Project modules
# from _profile.db import profiles


# Main script
def run():
    # parse(args)     Dont turn this on until the parse function is accurate.
    print("File 'db.py' loaded and run() started.")
    profiles()
    print("File 'db.py' loaded and run() completed.")


def handler(wb, conf: dict, app_models: list, base) -> None:
    # Prepare and respond to conversion
    data = {}
    for model, keys in conf.items():
        data.update({model: from_xlsx(wb[model], keys)})
    data.update({"Timestamp": dt.now().strftime("%Y-%m-%d %H-%M-%S.%f %Z")})
    t = data["Timestamp"]
    print(f"xlsx read {t}")
    load = to_json_file(data, t, base)
    print("json written")
    to_sql(load, app_models)
    print("sql saved")


def from_xlsx(ws, keys, active=True):
    # Convert from xlsx to json
    list = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        dictionary = {}
        for h in keys:
            dictionary[h] = row[keys.index(h)]
        dictionary["active"] = active
        list.append(dictionary)
    return list


def to_json_file(data, t, base):
    load = json.dumps(data)
    # Save to long-term archive
    with open(f"{base}data/db {t}.json", mode="w") as f:
        f.write(load)

    # Write to active database verifier
    with open(f"{base}db.json", mode="w") as f:
        f.write(load)

    return load


def to_sql(load, app_models):
    print(load)
    print(app_models)


def backup():
    which_list()

    def which_list():
        today = dt.datetime.now()
        year = today.year
        month = today.month
        weekday = today.weekday
        day = today.day
        # if month == 1 and day == 1:
        #     bak_annual(year)
        # if day == 1 and (month == 4 or month == 7 or month == 10):
        #     bak_quarter(today)
        # if day == 1:
        #     bak_monthly(today)
        # if weekday == 4:
        #     bak_weekly(today)
        # else:
        bak_daily(today)

    def bak_daily():
        pass
        # Write a script to create a daily backup, erasing backup older than 3 days.

    def bak_weekly():
        pass
        # Write a script to create a weekly backup, erasing backup older than 3 weeks.

    def bak_monthly():
        pass
        # Write a script to create a weekly backup, erasing backup older than 3 weeks.


# def parse():
#     global_parser = argparse.ArgumentParser(
#         prog="spd",
#         description="Play 'Paper-Rock-Scissors' using a Command Line Interface.",
#         # usage="",
#         epilog="By Scott Paul Doolittle 2024",
#     )
#     subparsers = global_parser.add_subparsers(
#         title="subcommands",
#     )
#     prs_parser = subparsers.add_parser(
#         "prs", help="A class game of Paper Rock Scissors."
#     )
#     prs_parser.add_argument(dest="prs", help="A class game of Paper Rock Scissors.")
#     prs_parser.set_defaults(func=prs)

#     two_parser = subparsers.add_parser("two", help="The 'other' game.")
#     two_parser.add_argument(dest="two", help="The 'other' game.")
#     two_parser.set_defaults(func=two)

#     args = global_parser.parse_args()

#     print(args)

#     def prs():
#         pass

#     def two():
#         pass


def lands():
    from _land import models as _land_models

    a = _land_models.District.objects.all()
    b = _land_models.FloodZones.objects.all()
    c = _land_models.Jurisdiction.objects.all()
    d = _land_models.Parcel.objects.all()
    e = _land_models.CityStZip.objects.all()
    f = _land_models.SiteAddress.objects.all()
    app_models = [a, b, c, d, e, f]

    base = "_land/"
    wb = load_workbook(f"{base}db.xlsx")
    conf = {
        # key = model
        # value = list of fields for the model
        #
        # Requirements:
        #   values must match '''xlsx''' file order. (OK to vary name)
        #   values must match '''model''' field name. (OK to mix order)
        # These two requirements allow this function to utilize an Excel
        # file that does not match the models in column order or field naming.
        "district": [
            "id",
            "type",
            "name",
            "description",
            "address",
            "city_st_zip",
            "phone",
            "email",
            "website",
        ],
        "floodzones": ["id", "name", "description"],
        "jurisdiction": ["id", "name"],
        "parcel": [
            "id",
            "book",
            "page",
            "parcel",
            "full",
            "owner_name",
            "owner_address",
            "land_use_zone",
            "wui_sra",
            "wuil_lra",
            "wui_risk",
            "wui_regulations",
            "flood_a",
            "flood_ae",
            "flood_ao",
            "flood_x",
            "floodway",
            "jurisdiction",
            "districts",
            "parcels",
        ],
        "citystzip": ["id", "city", "state", "zip"],
        "site_address": [
            "id",
            "parcel",
            "number",
            "street",
            "city_st_zip",
            "geolocation",
        ],
    }
    handler(wb, conf, app_models, base)


def profiles():
    from _profile import models as _profile_models

    a = _profile_models.Agency.objects.all()
    b = _profile_models.Certification.objects.all()
    c = _profile_models.Department.objects.all()
    d = _profile_models.Division.objects.all()
    e = _profile_models.Profile.objects.all()
    app_models = [a, b, c, d, e]

    base = "_profile/"
    wb = load_workbook(f"{base}db.xlsx")
    conf = {
        # key = model
        # value = list of fields for the model
        #
        # Requirements:
        #   values must match '''xlsx''' file order. (OK to vary name)
        #   values must match '''model''' field name. (OK to mix order)
        # These two requirements allow this function to utilize an Excel
        # file that does not match the models in column order or field naming.
        "agency": ["id", "name", "full"],
        "certification": ["id", "agency", "agency_long", "license", "license_long"],
        "department": ["id", "name", "full"],
        "division": ["id", "prefix", "division", "full_division"],
        "profile": [
            "id",
            "user",
            "first",
            "last",
            "email",
            "phone",
            "address",
            "city",
            "state",
            "zip",
            "staff",
            "superadmin",
            "company",
            "agency",
            "department",
            "division",
            "reviewer",
            "inspector",
            "alt_contact",
            "alt_contact_email",
            "alt_contact_phone",
        ],
    }
    handler(wb, conf, app_models, base)


print("File 'db.py' loaded.")

if __name__ == "__main__":
    run()
