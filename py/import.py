import json
from django.contrib.auth.models import User
from django.core import serializers
from django.db import IntegrityError
import openpyxl

from fees.models import Account, FeeType, PaymentMethod
from inspections.models import InspectionType, InspectionResult
from locations.models import District, Jurisdiction
from records.models import Tag, Division, Status, Type, Number
from profiles.models import Profile, LicenseAgency, LicenseType, Agency, Department, Division
from reviews.models import ReviewType, ReviewStatus, CycleResult

##########################################################################
""" python manage.py runscript import """
##########################################################################
base = ".init/"
def run():
    fiscal()
    inspections()
    locations()
    profiles()
    reviews()
    records()
    # permits_bp()

##########################################################################
""" Fiscal """
##########################################################################
def fiscal():
    wb = openpyxl.load_workbook(f'{base}fiscal.xlsx')

    ws = wb["accounts"]
    for i in range(2, 39):
        a = Account()
        a.id = ws[f"A{i}"].value
        a.fund = ws[f"B{i}"].value
        a.fund_label = ws[f"C{i}"].value
        a.share = ws[f"D{i}"].value
        a.unit = ws[f"E{i}"].value
        a.unit_label = ws[f"F{i}"].value
        a.unit_description = ws[f"G{i}"].value
        a.cost_center = ws[f"H{i}"].value
        a.gl_account = ws[f"I{i}"].value
        a.cams = ws[f"J{i}"].value
        a.infor_activity = ws[f"K{i}"].value
        a.infor_account = ws[f"L{i}"].value
        a.ledger = ws[f"M{i}"].value
        a.save()
    
    ws = wb["fee_types"]
    # acct = Account.objects.all()
    for i in range(2, 139):
        a = FeeType()
        a.id = ws[f"A{i}"].value
        # a.fee_account = acct.get(id=ws[f"B{i}"].value)
        a.fee_group = ws[f"C{i}"].value
        a.fee_name = ws[f"D{i}"].value
        a.policy = ws[f"E{i}"].value
        a.authorization = ws[f"F{i}"].value
        a.adopted = ws[f"G{i}"].value
        a.revised = ws[f"H{i}"].value
        a.expires = ws[f"I{i}"].value
        a.tier_base_qty = ws[f"J{i}"].value
        a.tier_base_fee = ws[f"K{i}"].value
        a.rate = ws[f"L{i}"].value
        a.units = ws[f"M{i}"].value
        a.rate_check = ws[f"N{i}"].value
        a.active = True
        a.deleted = False
        a.save()

    ws = wb["payment_methods"]
    for i in range(2, 8):
        a = PaymentMethod()
        a.id = ws[f"A{i}"].value
        a.method = ws[f"B{i}"].value
        a.policy = ws[f"C{i}"].value
        a.save()

##########################################################################
""" Locations """
##########################################################################
def locations():
    wb = openpyxl.load_workbook(f'{base}locations.xlsx')
    ws = wb["districts"]
    for i in range(2, 46):
        a = District()
        a.id = ws[f"A{i}"].value
        a.district_type = ws[f"B{i}"].value
        a.district = ws[f"C{i}"].value
        a.description = ws[f"D{i}"].value
        a.save()

    ws = wb["jurisdictions"]
    for i in range(2, 6):
        a = Jurisdiction()
        a.id = ws[f"A{i}"].value
        a.jurisdiction = ws[f"B{i}"].value
        a.save()

##########################################################################
""" Inspections """
##########################################################################
def inspections():
    wb = openpyxl.load_workbook(f'{base}inspections.xlsx')

    ws = wb["insp_type"]
    for i in range(2, 68):
        a = InspectionType()
        a.id = ws[f"A{i}"].value
        a.inspection_type = ws[f"B{i}"].value
        a.save()

    ws = wb["insp_result"]
    for i in range(2, 8):
        a = InspectionResult()
        a.id = ws[f"A{i}"].value
        a.result = ws[f"B{i}"].value
        a.requirements = ws[f"C{i}"].value
        a.save()

##########################################################################
""" Profiles """
##########################################################################
def profiles():
    wb = openpyxl.load_workbook(f'{base}profiles.xlsx')

    ws = wb["agency"]
    for i in range(2, 14):
        a = Agency()
        a.id = ws[f"A{i}"].value
        a.agency = ws[f"B{i}"].value
        a.full_agency = ws[f"C{i}"].value
        a.save()
    
    ws = wb["department"]
    for i in range(2, 18):
        a = Department()
        a.id = ws[f"A{i}"].value
        a.dept_code = ws[f"B{i}"].value
        a.department = ws[f"C{i}"].value
        a.save()
    
    ws = wb["division"]
    for i in range(2, 9):
        a = Division()
        a.id = ws[f"A{i}"].value
        a.prefix = ws[f"B{i}"].value
        a.division = ws[f"C{i}"].value
        a.full_division = ws[f"D{i}"].value
        a.save()
    
    ws = wb["profile"]
    for i in range(2, 31):
        a = User()
        a.id = ws[f"A{i}"].value
        a.username = ws[f"B{i}"].value
        a.first_name = ws[f"C{i}"].value
        a.last_name = ws[f"D{i}"].value
        a.email = ws[f"F{i}"].value
        a.is_active = ws[f"L{i}"].value
        a.is_staff = ws[f"M{i}"].value
        a.is_superuser = ws[f"N{i}"].value
        a.save()

    ws = wb["profile"]
    # acct = User.objects.all()
    for i in range(2, 31):
        a = Profile()
        a.id = ws[f"A{i}"].value
        # a.user = acct.get(id=ws[f"B{i}"].value)
        a.first = ws[f"C{i}"].value
        a.last = ws[f"D{i}"].value
        a.company = ws[f"E{i}"].value
        a.email = ws[f"F{i}"].value
        a.phone = ws[f"G{i}"].value
        a.address = ws[f"H{i}"].value
        a.city = ws[f"I{i}"].value
        a.state = ws[f"J{i}"].value
        a.zip = ws[f"K{i}"].value
        a.save()
    
    ws = wb["licenser"]
    for i in range(2, 6):
        a = LicenseAgency()
        a.id = ws[f"A{i}"].value
        a.agency = ws[f"B{i}"].value
        a.agency_long = ws[f"C{i}"].value
        a.save()

    ws = wb["license_type"]
    for i in range(2, 54):
        a = LicenseType()
        a.id = ws[f"A{i}"].value
        a.license_short = ws[f"B{i}"].value
        a.license_long = ws[f"C{i}"].value
        a.save()

##########################################################################
""" Reviews """
##########################################################################
def reviews():
    wb = openpyxl.load_workbook(f'{base}reviews.xlsx')
    
    ws = wb["review_type"]
    for i in range(2, 29):
        a = ReviewType()
        a.id = ws[f"A{i}"].value
        # a.review_division = ws[f"B{i}"].value
        # a.default_reviewer = ws[f"C{i}"].value
        a.review_type = ws[f"B{i}"].value
        a.days_cycle1 = ws[f"C{i}"].value
        a.days_cycle2 = ws[f"D{i}"].value
        # a.review_fees = ws[f"G{i}"].value
        a.save()

    ws = wb["review_status"]
    for i in range(2, 6):
        a = ReviewStatus()
        a.id = ws[f"A{i}"].value
        a.status = ws[f"B{i}"].value
        a.description = ws[f"C{i}"].value
        a.color = ws[f"D{i}"].value
        a.save()


    ws = wb["cycle_result"]
    for i in range(2, 7):
        a = CycleResult()
        a.id = ws[f"A{i}"].value
        a.result = ws[f"B{i}"].value
        a.description = ws[f"C{i}"].value
        a.causes_review_status = ws[f"D{i}"].value
        a.save()

##########################################################################
""" Permits """
##########################################################################
def records():
    wb = openpyxl.load_workbook(f'{base}permits.xlsx')
    
    ws = wb["sequence"]
    for i in range(2, 7):
        a = Number()
        a.id = ws[f"A{i}"].value
        a.series_title = ws[f"B{i}"].value
        a.series_prefix = ws[f"C{i}"].value
        a.year = ws[f"D{i}"].value
        a.sequence = ws[f"E{i}"].value
        a.save()

    ws = wb["tags"]
    for i in range(2, 10):
        a = Tag()
        a.id = ws[f"A{i}"].value
        a.tag = ws[f"B{i}"].value
        a.policy = ws[f"C{i}"].value
        a.save()

    ws = wb["permit_status"]
    for i in range(2, 13):
        a = Status()
        a.id = ws[f"A{i}"].value
        a.status = ws[f"B{i}"].value
        a.policy = ws[f"C{i}"].value
        a.build = ws[f"D{i}"].value
        a.occupy = ws[f"E{i}"].value
        a.save()

    ws = wb["permit_type"]
    for i in range(2, 18):
        a = Type()
        a.id = ws[f"A{i}"].value
        a.type = ws[f"B{i}"].value
        a.policy = ws[f"C{i}"].value
        a.suffix = ws[f"D{i}"].value
        a.save()


if __name__ == '__main__':
    run()
##########################################################################
""" End File """
##########################################################################